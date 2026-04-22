import pandas as pd
import asyncio
from sqlalchemy import select
from database.session import AsyncSessionLocal
from database.models import OrderLink
from openpyxl import load_workbook


async def export_long():
    rows = []

    async with AsyncSessionLocal() as session:
        result = await session.execute(select(OrderLink))
        orders = result.scalars().all()

        print(f"DB orders: {len(orders)}")

        for order in orders:
            students = (order.students_raw or "").split(",")

            for s in students:
                fio = s.strip()

                if not fio:
                    continue

                rows.append({
                    "FIO": fio,
                    "Buyruq turi": order.type,
                    "Buyruq nomi": order.title,
                    "Link": order.link
                })

    df = pd.DataFrame(rows)

    file_path = "files/orders_long.xlsx"
    df.to_excel(file_path, index=False)

    # 🔥 OPENPYXL bilan hyperlink qo‘shamiz
    wb = load_workbook(file_path)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        link = ws[f"D{row}"].value  # Link ustuni
        cell = ws[f"C{row}"]        # Buyruq nomi

        if link:
            cell.hyperlink = link
            cell.style = "Hyperlink"

    # 🔥 Link ustunini yashiramiz (optional)
    ws.column_dimensions["D"].hidden = True

    wb.save(file_path)

    print("✅ Interaktiv Excel tayyor!")


if __name__ == "__main__":
    asyncio.run(export_long())
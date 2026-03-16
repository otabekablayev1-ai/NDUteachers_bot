#heads.py to'liq
import os
import openpyxl
from aiogram import Router, F
from aiogram.types import (
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    FSInputFile,
    Message,
    CallbackQuery
)
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
from aiogram.types import BufferedInputFile
from database.db import (get_university_statistics, get_question_by_id,
)
from aiogram.types import BufferedInputFile
from aiogram.exceptions import TelegramBadRequest
from aiogram.fsm.context import FSMContext
from database.db import get_manager_rating_table
from aiogram.fsm.state import StatesGroup, State
from data.config import MANAGERS_BY_FACULTY, RAHBARLAR
from database.db import (
    get_latest_questions_for_manager,
    save_answer,
    mark_question_answered,
    save_manager_rating,
    user_already_rated,
    get_all_teachers,
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from aiogram.types import BufferedInputFile
from datetime import datetime
from database.db import get_all_students

router = Router()

class ReplyFSM(StatesGroup):
    waiting = State()   # 🔴 MUHIM

# =========================
#   FSM HOLATLARI
# =========================
class ReplyQuestionFSM(StatesGroup):  # DB orqali savol
    waiting = State()


class ReplyDirectFSM(StatesGroup):    # bevosita userga javob
    waiting = State()


class SendMSG(StatesGroup):
    # umumiy
    role = State()         # teacher / tutor / student / all
    faculty = State()
    department = State()
    fio = State()

    # talabalar uchun
    edu_type = State()     # bak / mag / all
    edu_form = State()     # Kunduzgi / Kechki / Sirtqi / ...
    stu_faculty = State()
    course = State()       # 1..5 / all
    group = State()
    student_fio = State()

    msg = State()          # yakuniy xabar


def is_faculty_manager(manager_id: int) -> bool:
    for fac in MANAGERS_BY_FACULTY.values():
        if manager_id in fac["teacher"] or manager_id in fac["student"]:
            return True
    return False

# =========================
#   /rahbar – faqat rahbarlar
# =========================
def get_global_managers():
    ids = []
    for lst in RAHBARLAR.values():
        ids.extend(lst)
    return list(set(ids))  # dublikatlar bo‘lmasin

def get_faculty_manager(role: str, faculty: str):
    """
    Talaba → student manager
    O‘qituvchi yoki Tyutor → teacher manager
    """
    fac = MANAGERS_BY_FACULTY.get(faculty)
    if not fac:
        return []

    if role == "Talaba":
        return fac.get("student", [])
    else:
        return fac.get("teacher", [])

# =========================
#   1) SAVOLLARNI KO‘RISH
# =========================
@router.message(F.text == "📥 Savollarni ko‘rish")
async def view_questions(message: Message):
    questions = await get_latest_questions_for_manager(
        manager_id=message.from_user.id
    )

    if not questions:
        await message.answer("📭 Siz uchun yangi savollar yo‘q.")
        return

    for q in questions:
        answered = q["answered"]
        status = "✅ <b>Javob berilgan</b>" if answered else "⚠️ <b>Javob kutilmoqda</b>"

        text = (
            f"📩 <b>Yangi savol</b>\n\n"
            f"👤 <b>F.I.Sh:</b> {q['fio']}\n"
            f"🏫 <b>Fakultet:</b> {q['faculty']}\n"
            f"🕓 <b>Vaqt:</b> {q['created_at']}\n\n"
            f"❓ <b>Savol:</b>\n{q['message_text']}\n\n"
            f"{status}"
        )

        kb = None
        if not answered:
            kb = InlineKeyboardMarkup(
                inline_keyboard=[[
                    InlineKeyboardButton(
                        text="✉️ Javob yozish",
                        callback_data=f"reply_{q['id']}"
                    )
                ]]
            )

        await message.answer(text, parse_mode="HTML", reply_markup=kb)
# =========================
#   JAVOB YOZISH
# =========================
@router.callback_query(F.data.startswith("reply_"))
async def start_reply(call: CallbackQuery, state: FSMContext):
    try:
        question_id = int(call.data.split("_")[1])

        q = await get_question_by_id(question_id)  # 🔥 await qo‘shildi

        if not q:
            await call.answer("❗ Savol topilmadi", show_alert=True)
            return

        await state.update_data(
            question_id=question_id,
            sender_id=q.sender_id
        )

        await call.message.answer(
            f"✏️ <b>{q.fio}</b> ga javob yozing:",
            parse_mode="HTML"
        )

        await state.set_state(ReplyFSM.waiting)
        await call.answer()

    except Exception as e:
        print("REPLY ERROR:", e)
        await call.answer("❌ Xatolik yuz berdi", show_alert=True)

@router.message(ReplyFSM.waiting, F.text | F.photo | F.document | F.video)
async def send_reply(message: Message, state: FSMContext):
    data = await state.get_data()

    question_id = data.get("question_id")
    sender_id = data.get("sender_id")
    manager_id = message.from_user.id

    if not question_id or not sender_id:
        await message.answer("❗ Xatolik: savol topilmadi.")
        await state.clear()
        return

    # 🔥 MUHIM — await qo‘shildi
    q = await get_question_by_id(question_id)
    faculty = q.faculty if q and q.faculty else "Noma’lum"

    header = (
        f"📬 <b>Sizning savolingizga javob</b>\n\n"
        f"🏫 Fakultet: <b>{faculty}</b>\n"
        f"👤 Rahbar: {message.from_user.full_name}\n"
        f"────────────────\n\n"
    )

    # 1️⃣ USERGA JAVOB
    if message.text:
        await message.bot.send_message(
            sender_id,
            header + message.text,
            parse_mode="HTML"
        )
        answer_text = message.text
    elif message.document:
        await message.bot.send_document(
            sender_id,
            message.document.file_id,
            caption=header
        )
        answer_text = "Hujjat"
    elif message.photo:
        await message.bot.send_photo(
            sender_id,
            message.photo[-1].file_id,
            caption=header
        )
        answer_text = "Rasm"
    elif message.video:
        await message.bot.send_video(
            sender_id,
            message.video.file_id,
            caption=header
        )
        answer_text = "Video"
    else:
        await message.answer("❗ Noma’lum format.")
        await state.clear()
        return

    # 2️⃣ DB ga yozish
    try:
        await save_answer(question_id, manager_id, answer_text)
        await mark_question_answered(question_id)
    except Exception as e:
        print("[HEADS] DB error:", e)

    # 3️⃣ Baholash tugmalari
    stars_kb = InlineKeyboardMarkup(
        inline_keyboard=[[
            InlineKeyboardButton(text="1", callback_data=f"rate:{question_id}:{manager_id}:1"),
            InlineKeyboardButton(text="2", callback_data=f"rate:{question_id}:{manager_id}:2"),
            InlineKeyboardButton(text="3", callback_data=f"rate:{question_id}:{manager_id}:3"),
            InlineKeyboardButton(text="4", callback_data=f"rate:{question_id}:{manager_id}:4"),
            InlineKeyboardButton(text="5", callback_data=f"rate:{question_id}:{manager_id}:5"),
        ]]
    )

    try:
        await message.bot.send_message(
            sender_id,
            "👍 <b>Javobni baholang:</b>",
            reply_markup=stars_kb,
            parse_mode="HTML"
        )
    except Exception as e:
        print("[HEADS] Rating yuborishda xato:", e)

    await message.answer("✅ Javob foydalanuvchiga yuborildi.")
    await state.clear()

## =========================
#   JAVOBGA BAHO QO‘YISH
# =========================
@router.callback_query(F.data.startswith("rate:"))
async def handle_rating(call: CallbackQuery):
    _, qid, manager_id, rating = call.data.split(":")
    question_id = int(qid)
    manager_id = int(manager_id)
    rating = int(rating)

    if await user_already_rated(call.from_user.id, manager_id, question_id):
        await call.answer("❗ Siz allaqachon baholagansiz", show_alert=True)
        return

    await save_manager_rating(
        teacher_id=call.from_user.id,
        manager_id=manager_id,
        question_id=question_id,
        rating=rating
    )

    await call.message.edit_reply_markup(reply_markup=None)
    await call.answer("⭐ Bahoyingiz qabul qilindi!", show_alert=True)

    await call.bot.send_message(
        manager_id,
        f"📊 Javobingizga ⭐ {rating} ball berildi"
    )

from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
from datetime import datetime
import textwrap


async def generate_manager_rating_image(rows, bot):

    width = 2800
    padding_x = 80
    padding_y = 80

    TITLE_SIZE_MAIN = 48
    TITLE_SIZE_SUB = 40
    HEADER_SIZE = 30
    FONT_SIZE = 32
    SMALL_SIZE = 22

    row_height = 100
    header_height = 170

    height = (
        padding_y * 2
        + 150
        + header_height
        + row_height * (len(rows) + 1)
        + 150
    )

    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)

    def load_font(size):
        try:
            return ImageFont.truetype("DejaVuSans.ttf", size)
        except:
            return ImageFont.load_default()

    font_title_main = load_font(TITLE_SIZE_MAIN)
    font_title_sub = load_font(TITLE_SIZE_SUB)
    font_header = load_font(HEADER_SIZE)
    font = load_font(FONT_SIZE)
    font_small = load_font(SMALL_SIZE)

    y = padding_y

    # ================= TITLE =================

    title1 = "Navoiy davlat universiteti Registrator ofisi"
    title2 = "Telegram @NDUteachers_bot ga kelib tushgan murojaatlarni ko‘rib chiqish"
    title3 = "samaradorligi bo‘yicha mas’ul ijrochilar faoliyati monitoringi"

    draw.text((width//2, y), title1, font=font_title_main, fill="black", anchor="mm")
    y += 60
    draw.text((width//2, y), title2, font=font_title_sub, fill="black", anchor="mm")
    y += 60
    draw.text((width//2, y), title3, font=font_title_sub, fill="black", anchor="mm")

    y += 40

    # ================= TABLE =================

    col_no = 120
    col_name = 700
    col_position = 350
    col_rate = 200
    col_answered = 260
    col_unanswered = 260
    col_fac = 700

    table_left = padding_x
    table_top = y

    table_right = (
            table_left
            + col_no
            + col_name
            + col_position
            + col_rate
            + col_answered
            + col_unanswered
            + col_fac
    )
    draw.rectangle(
        [table_left, table_top, table_right, table_top + header_height],
        fill=(235,235,235),
        outline="black",
        width=2
    )

    x_no = table_left
    x_name = x_no + col_no
    x_position = x_name + col_name
    x_rate = x_position + col_position
    x_ok = x_rate + col_rate
    x_bad = x_ok + col_answered
    x_fac = x_bad + col_unanswered

    center_y = table_top + header_height//2

    draw.text((x_no+col_no//2, center_y), "№", font=font_header, fill="black", anchor="mm")
    draw.text((x_name + col_name // 2, center_y), "Mas’ul ijrochi", font=font_header, fill="black", anchor="mm")
    draw.text((x_position + col_position // 2, center_y), "Lavozimi", font=font_header, fill="black", anchor="mm")
    draw.text((x_rate+col_rate//2, center_y-20), "Faoliyat", font=font_header, fill="black", anchor="mm")
    draw.text((x_rate+col_rate//2, center_y+20), "reytingi", font=font_header, fill="black", anchor="mm")

    for i, txt in enumerate(["Ko‘rib chiqilgan","murojaatlar","soni"]):
        draw.text((x_ok+col_answered//2, table_top+40+i*45), txt, font=font_header, fill="black", anchor="mm")

    for i, txt in enumerate(["Ko‘rib chiqilmagan","murojaatlar","soni"]):
        draw.text((x_bad+col_unanswered//2, table_top+40+i*45), txt, font=font_header, fill="black", anchor="mm")

    draw.text((x_fac+col_fac//2, center_y-20), "Murojaat", font=font_header, fill="black", anchor="mm")
    draw.text((x_fac+col_fac//2, center_y+20), "yo‘nalishi", font=font_header, fill="black", anchor="mm")

    columns = [
        x_no,
        x_name,
        x_position,
        x_rate,
        x_ok,
        x_bad,
        x_fac,
        table_right
    ]

    total_rows = len(rows)+1

    for col in columns:
        draw.line((col, table_top, col, table_top+header_height+row_height*total_rows), fill="black", width=2)

    # ================= NAME CACHE (tezroq) =================

    name_cache = {}

    async def get_name(manager_id):
        if manager_id in name_cache:
            return name_cache[manager_id]

        try:
            chat = await bot.get_chat(manager_id)
            name_cache[manager_id] = chat.full_name
        except:
            name_cache[manager_id] = str(manager_id)

        return name_cache[manager_id]

    # ================= DATA =================

    y = table_top + header_height

    total_answered = sum(int(r.get("answered_count",0)) for r in rows)
    total_unanswered = sum(int(r.get("unanswered_count",0)) for r in rows)

    medals = {1:"🥇",2:"🥈",3:"🥉"}

    for idx,r in enumerate(rows,1):

        row_bottom = y+row_height
        center_y = y+row_height//2

        if idx==1:
            draw.rectangle([table_left,y,table_right,row_bottom],fill=(255,248,220))

        name = await get_name(r["manager_id"])

        avg = float(r.get("avg_rating") or 0)

        medal = medals.get(idx,str(idx))

        draw.text((x_no + col_no // 2, center_y), medal, font=font, fill="black", anchor="mm")

        draw.text(
            (x_name + col_name // 2, center_y),
            name,
            font=font,
            fill="black",
            anchor="mm"
        )

        position = str(r.get("position", ""))

        draw.text(
            (x_position + col_position // 2, center_y),
            position,
            font=font,
            fill="black",
            anchor="mm"
        )


        draw.text((x_rate+col_rate//2,center_y),f"{avg:.1f}",font=font,fill="black",anchor="mm")

        draw.text((x_ok+col_answered//2,center_y),str(r.get("answered_count",0)),font=font,fill="black",anchor="mm")

        draw.text((x_bad+col_unanswered//2,center_y),str(r.get("unanswered_count",0)),font=font,fill="black",anchor="mm")

        faculty = str(r.get("faculty",""))

        draw.text(
            (x_fac + col_fac // 2, center_y),
            faculty,
            font=font,
            fill="black",
            anchor="mm"
        )

        draw.text(
            (x_fac + col_fac // 2, center_y),
            faculty,
            font=font,
            fill="black",
            anchor="mm"
        )
        draw.line((table_left,y,table_right,y),fill="black",width=2)
        draw.line((table_left,row_bottom,table_right,row_bottom),fill="black",width=2)

        for col in columns:
            draw.line((col,y,col,row_bottom),fill="black",width=2)

        y=row_bottom

    # ================= TOTAL =================

    center_y = y+row_height//2

    draw.text((x_name+20,center_y-20),"Jami murojaatlar soni:",font=font_header,fill="black")

    draw.text((x_ok+col_answered//2,center_y),str(total_answered),font=font_header,fill="black",anchor="mm")

    draw.text((x_bad+col_unanswered//2,center_y),str(total_unanswered),font=font_header,fill="black",anchor="mm")

    # ================= ECONOMY =================

    saved_money = total_answered*5000

    info_text = (
        "Telegram bot orqali qabul qilingan murojaatlar natijasida "
        "iqtisod qilingan transport xarajatlari (1 tashrif — 5 000 so‘m bo‘lganda)"
    )

    info_y = y+150

    center_x = table_left+(table_right-table_left)//2

    text_width = draw.textlength(info_text,font=font_header)

    draw.text((center_x-text_width//2,info_y),info_text,font=font_header,fill=(0,128,0))

    sum_text = f"{saved_money:,} so‘m"

    draw.text((table_right-150,info_y+40),sum_text,font=font_header,fill=(0,128,0),anchor="mm")

    # ================= FOOTER =================

    footer = f"Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    draw.text((padding_x,height-padding_y),footer,font=font_small,fill="black")

    buffer = BytesIO()
    img.save(buffer,format="PNG")
    buffer.seek(0)

    return buffer

@router.message(F.text == "🏆 Menejerlar reytingi")
async def show_managers_rating(message: Message):

    rows = await get_manager_rating_table()

    if not rows:
        await message.answer("📭 Hozircha menejerlar reytingi mavjud emas.")
        return

    image_buffer = await generate_manager_rating_image(rows, message.bot)

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(
                text="📤 Excelga eksport",
                callback_data="export_manager_rating_excel"
            )]
        ]
    )

    image_buffer = await generate_manager_rating_image(rows, message.bot)

    await message.answer_photo(
        photo=BufferedInputFile(image_buffer.read(), filename="menejer_reyting.png"),
        caption="📊 Reyting jadvali"
    )

@router.callback_query(F.data == "export_manager_rating_excel")
async def export_manager_rating_excel(call: CallbackQuery):

    rows = await get_manager_rating_table()

    if not rows:
        await call.answer("Ma'lumot topilmadi", show_alert=True)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menejerlar reytingi"

    ws.append([
        "T/r",
        "Menejer",
        "Reyting",
        "Javob berilgan",
        "Javob berilmagan",
        "Fakultet"
    ])

    for i, r in enumerate(rows, 1):
        try:
            chat = await call.bot.get_chat(r["manager_id"])
            name = chat.full_name
        except TelegramBadRequest:
            name = str(r["manager_id"])

        ws.append([
            i,
            name,
            r["avg_rating"],
            r["answered_count"],
            r["unanswered_count"],

            r["faculty"]
        ])

    path = "menejerlar_reytingi.xlsx"
    wb.save(path)

    await call.message.answer_document(
        FSInputFile(path),
        caption="📊 Menejerlar reytingi (Excel)"
    )

    await call.answer()

    if os.path.exists(path):
        os.remove(path)
# ==============================
#   📊 UNIVERSITET SUPER STATISTIKASI
# ==============================
@router.message(lambda m: m.text and "Statistika" in m.text)
async def full_stat(message: Message):

    stats = await get_university_statistics()

    text = (
        "<b>📊 UNIVERSITET UMUMIY STATISTIKASI</b>\n\n"
        f"👥 <b>Umumiy foydalanuvchilar:</b> {stats['total_users']} ta\n"
        f"👨‍🏫 <b>O‘qituvchilar:</b> {stats['teacher_count']} ta\n"
        f"🧑‍🏫 <b>Tyutorlar:</b> {stats['tutor_count']} ta\n"
        f"🎓 <b>Talabalar:</b> {stats['student_count']} ta\n\n"
        "<b>🏫 Fakultetlar bo‘yicha:</b>\n"
    )

    for fac, cnt in sorted(stats["faculty_stat"].items()):
        text += f"• {fac}: {cnt} ta\n"

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="📥 Excelga yuklab olish",
                    callback_data="export_stats_excel"
                )
            ]
        ]
    )

    await message.answer(text, parse_mode="HTML", reply_markup=kb)


@router.callback_query(F.data == "export_stats_excel")
async def export_stats_excel(call: CallbackQuery):
    await call.answer()   # 🔥 birinchi qatorda

    print("EXPORT CALLBACK IS WORKING")

    stats = await get_university_statistics()
    teachers = await get_all_teachers()
    students = await get_all_students()

    def pick(obj, key: str, default=""):
        # obj dict bo'lsa
        if isinstance(obj, dict):
            return obj.get(key, default)
        # obj model bo'lsa
        return getattr(obj, key, default)

    wb = Workbook()

    # =========================
    # 1-SHEET: STATISTIKA
    # =========================
    ws1 = wb.active
    ws1.title = "Statistika"

    ws1["A1"] = "UNIVERSITET UMUMIY STATISTIKASI"
    ws1.merge_cells("A1:B1")
    ws1["A1"].font = Font(size=14, bold=True)

    ws1.append(["Ko‘rsatkich", "Qiymat"])
    ws1.append(["Umumiy foydalanuvchilar", stats["total_users"]])
    ws1.append(["O‘qituvchilar", stats["teacher_count"]])
    ws1.append(["Tyutorlar", stats["tutor_count"]])
    ws1.append(["Talabalar", stats["student_count"]])

    ws1.append([])
    ws1.append(["Fakultetlar bo‘yicha"])

    for fac, cnt in stats["faculty_stat"].items():
        ws1.append([fac, cnt])

    # =========================
    # 2-SHEET: FOYDALANUVCHILAR
    # =========================
    ws2 = wb.create_sheet("Foydalanuvchilar")

    headers = [
        "Telegram ID",
        "F.I.O",
        "Telefon",
        "Rol",
        "Fakultet",
        "Ta'lim turi",
        "Ta'lim shakli",
        "Kurs",
        "Guruh",
        "Ro‘yxatdan o‘tgan sana"
    ]

    ws2.append(headers)

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    # Teachers & Tutors
    for t in teachers:
        role_val = pick(t, "role", "")
        role_label = "O‘qituvchi" if role_val == "teacher" else ("Tyutor" if role_val == "tutor" else role_val)

        created = pick(t, "created_at", None)
        ws2.append([
            pick(t, "user_id"),
            pick(t, "fio"),
            pick(t, "phone", ""),
            role_label,
            pick(t, "faculty", ""),
            "", "", "", "",  # edu_type, edu_form, course, group teacherda yo'q
            created.strftime("%Y-%m-%d") if created else ""
        ])

    # Students
    for s in students:
        created = pick(s, "created_at", None)
        ws2.append([
            pick(s, "user_id"),
            pick(s, "fio"),
            pick(s, "phone", ""),
            "Talaba",
            pick(s, "faculty", ""),
            pick(s, "edu_type", ""),
            pick(s, "edu_form", ""),
            pick(s, "course", ""),
            pick(s, "student_group", ""),
            created.strftime("%Y-%m-%d") if created else ""
        ])
    # Auto column width
    for column in ws2.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max_length + 2

    # =========================
    # Faylni saqlash
    # =========================
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"universitet_statistika_{today}.xlsx"

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    await call.message.answer_document(
        BufferedInputFile(file_stream.read(), filename=filename)
    )

